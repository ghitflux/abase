from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.db.models import Sum, Count
from apps.cadastros.models import Cadastro
from apps.cadastros.choices import StatusCadastro


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = cell.value
            l = len(str(val)) if val else 0
            if l > max_len:
                max_len = l
        ws.column_dimensions[col_letter].width = min(max_len+2, 60)


def build_xlsx_resumo_associados(qs):
    wb = Workbook()
    ws = wb.active
    ws.title = "Associados"
    headers = [
        "ID", "Nome/Razão Social", "CPF/CNPJ", "Matrícula", "Órgão Público",
        "Status", "Contribuição (R$) x3", "Doação (R$)", "Disponível (R$)",
        "Criado em", "Aprovado em", "Efetivado em"
    ]
    ws.append(headers)
    for c in qs:
        doc = c.cpf or c.cnpj or "—"
        ws.append([
            c.id, c.nome_completo, doc, c.matricula_servidor, c.orgao_publico,
            c.get_status_display(), float(c.valor_total_antecipacao or 0),
            float(c.doacao_associado or 0), float(c.disponivel or 0),
            c.created_at and c.created_at.strftime("%d/%m/%Y %H:%M"),
            c.approved_at and c.approved_at.strftime("%d/%m/%Y %H:%M"),
            c.paid_at and c.paid_at.strftime("%d/%m/%Y %H:%M"),
        ])
    _auto_width(ws)
    return wb


def build_xlsx_resumo_por_orgao(qs):
    wb = Workbook()
    ws = wb.active
    ws.title = "Órgão Público (Síntese)"
    ws.append(["Órgão Público", "Qtde Assoc.", "Total Contribuições (R$)", "Doações (R$)"])
    agg = (qs.values("orgao_publico")
             .annotate(qtd=Count("id"),
                       total=Sum("valor_total_antecipacao"),
                       doacoes=Sum("doacao_associado"))
             .order_by("-qtd"))
    for row in agg:
        ws.append([
            row["orgao_publico"] or "—",
            row["qtd"],
            float(row["total"] or 0),
            float(row["doacoes"] or 0),
        ])
    _auto_width(ws)
    return wb