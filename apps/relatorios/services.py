from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.db.models import Sum, Count
from django.template.loader import render_to_string
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import inch
from apps.cadastros.models import Cadastro
from apps.cadastros.choices import StatusCadastro
import io
from datetime import datetime


# ===== XLSX EXPORT FUNCTIONS =====

def _auto_width(ws):
    """Ajusta automaticamente a largura das colunas da planilha."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = cell.value
            l = len(str(val)) if val else 0
            if l > max_len:
                max_len = l
        ws.column_dimensions[col_letter].width = min(max_len+2, 60)


def build_xlsx_resumo_associados(qs):
    """Gera planilha XLSX com resumo dos associados."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Associados"
    headers = [
        "ID", "Nome/Razão Social", "CPF/CNPJ", "Matrícula", "Órgão Público",
        "Status", "Contribuição (R$) x3", "Doação (R$)", "Disponível (R$)",
        "Criado em", "Aprovado em", "Efetivado em"
    ]
    ws.append(headers)
    for c in qs:
        doc = c.cpf or c.cnpj or "—"
        ws.append([
            c.id, c.nome_razao_social, doc, c.matricula_servidor, c.orgao_publico,
            c.get_status_display(), float(c.valor_total_antecipacao or 0),
            float(c.doacao_associado or 0), float(c.disponivel or 0),
            c.created_at and c.created_at.strftime("%d/%m/%Y %H:%M"),
            c.approved_at and c.approved_at.strftime("%d/%m/%Y %H:%M"),
            c.paid_at and c.paid_at.strftime("%d/%m/%Y %H:%M"),
        ])
    _auto_width(ws)
    return wb


def build_xlsx_resumo_por_orgao(qs):
    """Gera planilha XLSX com resumo por órgão público."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Órgão Público (Síntese)"
    ws.append(["Órgão Público", "Qtde Assoc.", "Total Contribuições (R$)", "Doações (R$)"])
    agg = (qs.values("orgao_publico")
             .annotate(qtd=Count("id"),
                       total=Sum("valor_total_antecipacao"),
                       doacoes=Sum("doacao_associado"))
             .order_by("-qtd"))
    for row in agg:
        ws.append([
            row["orgao_publico"] or "—",
            row["qtd"],
            float(row["total"] or 0),
            float(row["doacoes"] or 0),
        ])
    _auto_width(ws)
    return wb


# ===== PDF GENERATION FUNCTIONS =====

def generate_pdf_associados(qs, title="Relatório de Associados"):
    """Gera PDF com relatório de associados."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    
    # Criar estilos personalizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Centralizado
    )
    
    elements = []
    
    # Título
    title_para = Paragraph(title, title_style)
    elements.append(title_para)
    elements.append(Spacer(1, 12))
    
    # Resumo
    total_associados = qs.count()
    total_contribuicoes = qs.aggregate(Sum('valor_total_antecipacao'))['valor_total_antecipacao__sum'] or 0
    total_doacoes = qs.aggregate(Sum('doacao_associado'))['doacao_associado__sum'] or 0
    
    summary_text = f"""Total de Associados: {total_associados}<br/>
    Total de Contribuições: R$ {total_contribuicoes:,.2f}<br/>
    Total de Doações: R$ {total_doacoes:,.2f}"""
    
    summary_para = Paragraph(summary_text, styles['Normal'])
    elements.append(summary_para)
    elements.append(Spacer(1, 20))
    
    # Tabela de dados
    data = [['ID', 'Nome/Razão Social', 'CPF/CNPJ', 'Status', 'Contribuição (R$)']]
    
    for c in qs:
        doc_num = c.cpf or c.cnpj or "—"
        data.append([
            str(c.id),
            c.nome_razao_social,
            doc_num,
            c.get_status_display(),
            f"R$ {float(c.valor_total_antecipacao or 0):,.2f}"
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Rodapé
    footer_text = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    footer_para = Paragraph(footer_text, styles['Normal'])
    elements.append(Spacer(1, 20))
    elements.append(footer_para)
    
    doc.build(elements)
    pdf_content = buffer.getvalue()
    buffer.close()
    
    return pdf_content


def generate_pdf_por_orgao(qs, title="Relatório por Órgão Público"):
    """Gera PDF com relatório por órgão público."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    
    # Criar estilos personalizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Centralizado
    )
    
    elements = []
    
    # Título
    title_para = Paragraph(title, title_style)
    elements.append(title_para)
    elements.append(Spacer(1, 12))
    
    # Agregação de dados
    agg = (qs.values("orgao_publico")
             .annotate(qtd=Count("id"),
                       total=Sum("valor_total_antecipacao"),
                       doacoes=Sum("doacao_associado"))
             .order_by("-qtd"))
    
    # Resumo
    total_orgaos = agg.count()
    total_associados = qs.count()
    total_contribuicoes = qs.aggregate(Sum('valor_total_antecipacao'))['valor_total_antecipacao__sum'] or 0
    total_doacoes = qs.aggregate(Sum('doacao_associado'))['doacao_associado__sum'] or 0
    
    summary_text = f"""Total de Órgãos: {total_orgaos}<br/>
    Total de Associados: {total_associados}<br/>
    Total de Contribuições: R$ {total_contribuicoes:,.2f}<br/>
    Total de Doações: R$ {total_doacoes:,.2f}"""
    
    summary_para = Paragraph(summary_text, styles['Normal'])
    elements.append(summary_para)
    elements.append(Spacer(1, 20))
    
    # Tabela de dados
    data = [['Órgão Público', 'Qtde Assoc.', 'Total Contribuições (R$)', 'Doações (R$)']]
    
    for row in agg:
        data.append([
            row["orgao_publico"] or "—",
            str(row["qtd"]),
            f"R$ {float(row['total'] or 0):,.2f}",
            f"R$ {float(row['doacoes'] or 0):,.2f}"
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Rodapé
    footer_text = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    footer_para = Paragraph(footer_text, styles['Normal'])
    elements.append(Spacer(1, 20))
    elements.append(footer_para)
    
    doc.build(elements)
    pdf_content = buffer.getvalue()
    buffer.close()
    
    return pdf_content


def generate_dashboard_pdf(context, title="Dashboard Diretoria"):
    """Gera PDF do dashboard da diretoria."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    
    # Criar estilos personalizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Centralizado
    )
    
    metric_style = ParagraphStyle(
        'MetricStyle',
        parent=styles['Normal'],
        fontSize=14,
        spaceAfter=10,
        alignment=1
    )
    
    elements = []
    
    # Título
    title_para = Paragraph(title, title_style)
    elements.append(title_para)
    elements.append(Spacer(1, 12))
    
    # Métricas principais
    metrics_data = []
    if 'total_associados' in context:
        metrics_data.append(['Total Associados', str(context['total_associados'])])
    if 'total_contribuicoes' in context:
        metrics_data.append(['Total Contribuições', f"R$ {context['total_contribuicoes']:,.2f}"])
    if 'total_doacoes' in context:
        metrics_data.append(['Total Doações', f"R$ {context['total_doacoes']:,.2f}"])
    if 'total_disponivel' in context:
        metrics_data.append(['Total Disponível', f"R$ {context['total_disponivel']:,.2f}"])
    
    if metrics_data:
        metrics_table = Table(metrics_data, colWidths=[3*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(metrics_table)
        elements.append(Spacer(1, 20))
    
    # Tabelas adicionais do contexto
    if 'status_counts' in context:
        elements.append(Paragraph("Status dos Associados", styles['Heading2']))
        status_data = [['Status', 'Quantidade']]
        for status, count in context['status_counts'].items():
            status_data.append([status, str(count)])
        
        status_table = Table(status_data)
        status_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(status_table)
        elements.append(Spacer(1, 20))
    
    # Rodapé
    footer_text = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    footer_para = Paragraph(footer_text, styles['Normal'])
    elements.append(footer_para)
    
    doc.build(elements)
    pdf_content = buffer.getvalue()
    buffer.close()
    
    return pdf_content